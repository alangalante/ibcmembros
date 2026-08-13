#!/usr/bin/env python3
"""Gera uma prévia determinística da migração. Não acessa Firebase nem produção."""

from __future__ import annotations

import argparse, csv, hashlib, json, re, unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
import cv2

PARTICLES = {"da", "das", "de", "do", "dos", "e"}
ADMIN_NAMES = {"Alan Carvalho Galante", "Tarcísio Nunes Cardoso", "Ackley de Almeida Fontes", "Ruiter de Campos Muniz"}
SPECIAL_LEADERS = {
    "Grupo de Adolescentes": ["Ariana Pires de Almeida Medeiros", "Luana Cristine dos Santos Lima Ayres", "Lucas Medeiros Pereira Nascimento", "Rodrigo de Souza Ayres"],
    "Grupo de Jovens": ["Rafael Brasil de Aguiar", "Aline Alves Ramos Brasil"],
}

def ascii_key(value: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", value or "") if unicodedata.category(c) != "Mn").lower()

def pretty_name(value: str) -> str:
    words = re.sub(r"\s+", " ", value.strip()).lower().split()
    return " ".join(word if index and word in PARTICLES else word[:1].upper() + word[1:] for index, word in enumerate(words))

def username_base(name: str) -> str:
    words = [re.sub(r"[^a-z0-9]", "", ascii_key(word)) for word in name.split()]
    words = [word for word in words if word]
    if len(words) == 1: return f"{words[0]}.{words[0]}"
    return f"{words[0]}{words[1][0]}.{words[-1]}"

def initial_password(name: str) -> str:
    """Regra única e comunicável: último sobrenome normalizado + 123."""
    return f"{ascii_key(name.split()[-1])}123"

def normalize_phone(value: object) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if digits.startswith("55") and len(digits) in (12, 13): digits = digits[2:]
    return digits if len(digits) in (10, 11) else ""

def iso_date(value: object) -> str | None:
    if isinstance(value, (datetime, date)): return value.strftime("%Y-%m-%d")
    if not value: return None
    text = str(value).strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d"):
        try: return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError: pass
    return None

def group_info(raw: str) -> tuple[str | None, str | None]:
    raw = str(raw or "").strip()
    if not raw or ascii_key(raw) == "nenhum": return None, None
    numbered = re.match(r"^(\d+)\s*-\s*(.+)$", raw)
    if numbered: return f"Grupo {int(numbered.group(1))}", pretty_name(numbered.group(2))
    if ascii_key(raw) == "adolescentes": return "Grupo de Adolescentes", None
    if ascii_key(raw) == "jovens": return "Grupo de Jovens", None
    return pretty_name(raw), None

def photo_tokens(path: Path) -> set[str]:
    stem = re.sub(r"^\d{1,2}-\d{1,2}\s*", "", path.stem)
    return {token for token in re.findall(r"[a-z]{3,}", ascii_key(stem)) if token not in PARTICLES}

def likely_single(path: Path) -> bool:
    name = ascii_key(path.stem)
    return "," not in path.stem and not re.search(r"\b(e|com)\b|\+|famil|casal|grupo", name)

def image_metrics(path: Path, detector) -> dict | None:
    image = cv2.imread(str(path))
    if image is None: return None
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    height, width = gray.shape
    scale = min(1.0, 1400 / max(width, height))
    sample = cv2.resize(gray, None, fx=scale, fy=scale) if scale < 1 else gray
    detected = detector.detectMultiScale(sample, scaleFactor=1.08, minNeighbors=5, minSize=(35, 35))
    faces = sorted(detected, key=lambda face: face[2] * face[3], reverse=True)
    # O Haar encontra falsos positivos pequenos em roupas e bordas transparentes.
    # Só contam como outras pessoas rostos com pelo menos 25% da área do principal.
    if faces:
        largest_area = faces[0][2] * faces[0][3]
        faces = [face for face in faces if face[2] * face[3] >= largest_area * .25]
    result = {"faceCount": len(faces), "width": width, "height": height, "contentHash": hashlib.sha256(path.read_bytes()).hexdigest(), "faceArea": 0.0, "sharpness": 0.0, "centeredness": 0.0}
    if len(faces) == 1:
        x, y, w, h = faces[0]
        crop = sample[y:y+h, x:x+w]
        result["faceArea"] = float(w * h / (sample.shape[0] * sample.shape[1]))
        result["sharpness"] = float(cv2.Laplacian(crop, cv2.CV_64F).var())
        face_x, face_y = (x + w / 2) / sample.shape[1], (y + h / 2) / sample.shape[0]
        result["centeredness"] = max(0.0, 1 - ((face_x - .5) ** 2 + (face_y - .43) ** 2) ** .5)
    return result

def run(source: Path, photos: Path, output: Path) -> None:
    sheet = load_workbook(source, data_only=True, read_only=True).active
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=8, values_only=True), 8):
        name, nickname, birth, phone1, phone2, raw_group, status = row[1:8]
        if ascii_key(str(status or "")).upper() != "MEMBRO": continue
        rows.append({"sourceRow": row_number, "name": pretty_name(str(name)), "nickname": pretty_name(str(nickname)) if nickname else "", "birthDate": iso_date(birth), "phoneE164": normalize_phone(phone1), "secondaryPhone": normalize_phone(phone2), "rawGroup": str(raw_group or "")})

    used = Counter(); members = []
    for row in rows:
        base = username_base(row["name"]); used[base] += 1
        username = base if used[base] == 1 else f"{base}{used[base]}"
        group, partial_leader = group_info(row.pop("rawGroup"))
        members.append({**row, "username": username, "initialPassword": initial_password(row["name"]), "mustChangePassword": True, "group": group, "partialLeader": partial_leader, "role": "admin" if row["name"] in ADMIN_NAMES else "common", "photo": None})

    by_name = {ascii_key(member["name"]): member for member in members}
    group_leaders: dict[str, list[str]] = defaultdict(list)
    errors = []
    for member in members:
        if member["group"] and member["partialLeader"]:
            query = set(ascii_key(member["partialLeader"]).split())
            matches = [candidate for candidate in members if query <= set(ascii_key(candidate["name"]).split())]
            if len(matches) == 1: group_leaders[member["group"]].append(matches[0]["name"])
            elif len(matches) != 0: errors.append(f"Líder ambíguo: {member['partialLeader']}")
    for group, leaders in SPECIAL_LEADERS.items(): group_leaders[group].extend(leaders)
    for leaders in group_leaders.values():
        for name in leaders:
            if ascii_key(name) not in by_name: errors.append(f"Líder não encontrado: {name}")
            elif by_name[ascii_key(name)]["role"] != "admin": by_name[ascii_key(name)]["role"] = "leader"

    detector = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    metrics = {}
    for path in photos.rglob("*"):
        if path.is_file() and not path.name.startswith("."):
            details = image_metrics(path, detector)
            if details: metrics[str(path)] = details
    candidates: dict[str, list[dict]] = defaultdict(list)
    for path in photos.rglob("*"):
        if not path.is_file() or path.name.startswith(".") or not likely_single(path): continue
        tokens = photo_tokens(path)
        for member in members:
            name_tokens = set(ascii_key(member["name"]).split()) - PARTICLES
            first_name = ascii_key(member["name"].split()[0])
            nickname_tokens = set(ascii_key(member["nickname"]).split()) - PARTICLES if member["nickname"] else set()
            identity_match = first_name in tokens or bool(tokens & nickname_tokens)
            if not identity_match: continue
            overlap = tokens & name_tokens
            birth = member["birthDate"]
            dated = re.match(r"^(\d{1,2})-(\d{1,2})", path.stem)
            date_match = bool(birth and dated and int(dated.group(1)) == int(birth[8:10]) and int(dated.group(2)) == int(birth[5:7]))
            if (date_match and overlap) or len(overlap) >= 2:
                details = metrics.get(str(path))
                if not details: continue
                match_score = (3 if date_match else 0) + len(overlap) * 2
                quality_score = min(details["sharpness"] / 250, 4) + min(details["faceArea"] * 20, 3) + details["centeredness"]
                score = match_score + quality_score
                candidates[member["name"]].append({"path": str(path), "score": round(score, 2), "matchScore": match_score, **details})
    claimed_photos = set(); claimed_hashes = set()
    ranked_members = sorted(members, key=lambda member: max((item["score"] for item in candidates[member["name"]] if item["faceCount"] == 1 or item["matchScore"] >= 5), default=0), reverse=True)
    for member in ranked_members:
        options = sorted(candidates[member["name"]], key=lambda item: (-item["score"], -(item["width"] * item["height"]), item["path"]))
        safe = [option for option in options if (option["faceCount"] == 1 or option["matchScore"] >= 5) and option["path"] not in claimed_photos and option["contentHash"] not in claimed_hashes]
        if safe: member["photo"] = safe[0]["path"]
        if member["photo"]:
            claimed_photos.add(member["photo"]); claimed_hashes.add(safe[0]["contentHash"])

    groups = []
    for name in sorted({m["group"] for m in members if m["group"]}, key=lambda value: (not value.startswith("Grupo ") or not value[6:].isdigit(), int(value[6:]) if value[6:].isdigit() else value)):
        participants = [m["username"] for m in members if m["group"] == name]
        leaders = list(dict.fromkeys(group_leaders[name]))
        groups.append({"name": name, "participantUsernames": participants, "leaderNames": leaders})

    usernames = [m["username"] for m in members]
    if len(usernames) != len(set(usernames)): errors.append("Há nomes de usuário duplicados")
    output.mkdir(parents=True, exist_ok=True)
    (output / "migration-preview.json").write_text(json.dumps({"members": members, "groups": groups}, ensure_ascii=False, indent=2), encoding="utf-8")
    with (output / "members-preview.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["sourceRow", "name", "username", "initialPassword", "birthDate", "phoneE164", "secondaryPhone", "group", "role", "photo"]
        writer = csv.DictWriter(handle, fields); writer.writeheader(); writer.writerows(({key: m.get(key) or "" for key in fields} for m in members))
    with (output / "photo-review.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["name", "status", "selectedPhoto", "candidateCount", "singleFaceCandidates", "selectedScore", "faceArea", "sharpness"]
        writer = csv.DictWriter(handle, fields); writer.writeheader()
        for member in members:
            options = candidates[member["name"]]; selected = next((item for item in options if item["path"] == member["photo"]), {})
            writer.writerow({"name": member["name"], "status": "SELECIONADA_REVISAR" if member["photo"] else "SEM_FOTO_SEGURA", "selectedPhoto": member["photo"] or "", "candidateCount": len(options), "singleFaceCandidates": sum(item["faceCount"] == 1 for item in options), "selectedScore": selected.get("score", ""), "faceArea": round(selected.get("faceArea", 0), 4) if selected else "", "sharpness": round(selected.get("sharpness", 0), 1) if selected else ""})
    without_photo = sorted((member for member in members if not member["photo"]), key=lambda member: ((member["birthDate"] or "9999-99-99")[5:], member["name"]))
    with (output / "membros-sem-foto.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        fields = ["Nome", "Data de nascimento", "Dia", "Mês", "Usuário"]
        writer = csv.DictWriter(handle, fields); writer.writeheader()
        for member in without_photo:
            birth = member["birthDate"] or ""
            writer.writerow({"Nome": member["name"], "Data de nascimento": birth, "Dia": birth[8:10] if birth else "", "Mês": birth[5:7] if birth else "", "Usuário": member["username"]})
    report = {"members": len(members), "groups": len(groups), "admins": sum(m["role"] == "admin" for m in members), "leaders": sum(m["role"] == "leader" for m in members), "withoutBirthDate": sum(not m["birthDate"] for m in members), "withoutPhone": sum(not m["phoneE164"] for m in members), "initialPasswordRule": "ultimo_sobrenome_sem_acento_em_minusculas+123", "photoFilesAnalyzed": len(metrics), "filesWithOneDetectedFace": sum(item["faceCount"] == 1 for item in metrics.values()), "filesWithMultipleDetectedFaces": sum(item["faceCount"] > 1 for item in metrics.values()), "filesWithoutDetectedFace": sum(item["faceCount"] == 0 for item in metrics.values()), "withSelectedPhoto": sum(bool(m["photo"]) for m in members), "photoSelectionNeedsHumanReview": True, "errors": errors}
    (output / "validation-report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if errors: raise SystemExit(2)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=Path("data/MEMBROS - ALAN.xlsx"))
    parser.add_argument("--photos", type=Path, default=Path("data/Fotos"))
    parser.add_argument("--output", type=Path, default=Path("migration-output"))
    args = parser.parse_args(); run(args.source, args.photos, args.output)
